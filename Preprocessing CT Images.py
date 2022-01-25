'''
Dataloader: Preprocessing existing CT and X-ray images into the h5 file format for the X2CT GAN Network

Indexes: f: frontal, l: lateral, right/left Knee
KOS: x-y in flat plane, z in height
'''

#Import libraries
import os
from pydicom import dcmread
from openpyxl import load_workbook
from skimage.transform import resize
import numpy as np
import h5py
from skimage.util import img_as_ubyte

def cutImg(list, x,y, pxl):
    pxl=int(pxl)
    if pxl % 2 == 1:
        pxl=pxl+1
    pxl=int(pxl/2) # cut evenly in all directions with length pxl/2 from the reference point
    paddingarr=np.zeros((2*pxl,2*pxl),  dtype= list.dtype) 
    #Cut image outside left boundary
    if x-pxl<0: 
        if y-pxl<0: #boundary outside top
            paddingarr[ abs(y-pxl):pxl*2, abs(x-pxl): pxl*2]= list[0 : pxl*2-abs(y-pxl), 0 : pxl*2-abs(x-pxl)]
        elif y+pxl> list.shape[0]: #boundary outside bottom
            paddingarr[ 0:2*pxl-(y+pxl-list.shape[0]), abs(x-pxl): pxl*2]= list[y-pxl : list.shape[0], 0 : pxl*2-abs(x-pxl)]
        else: #within height boundaries
            paddingarr[0:2*pxl, abs(x-pxl): pxl*2]= list[y-pxl: y+pxl, 0 : pxl*2-abs(x-pxl)]
    #Cut image outside right boundary
    elif x+pxl> list.shape[1]:
        if y-pxl<0: 
            paddingarr[ abs(y-pxl):pxl*2, 0: 2*pxl-(x+pxl-list.shape[1])]= list[0 :pxl*2-abs(y-pxl), x-pxl : list.shape[1]] #abs(x+pxl-list.shape[1])]
        elif y+pxl> list.shape[0]: 
            paddingarr[ 0:2*pxl-(y+pxl-list.shape[0]), 0: 2*pxl-(x+pxl-list.shape[1])]= list[y-pxl:list.shape[0],  x-pxl : list.shape[1]]
        else: 
            paddingarr[0:2*pxl, 0: 2*pxl-(x+pxl-list.shape[1])]= list[y-pxl: y+pxl, x-pxl : list.shape[1]]
    #Within horizontal boundaries
    else:
        if y-pxl<0: 
            paddingarr[ abs(y-pxl):pxl*2, 0: 2*pxl]= list[0 :pxl*2-abs(y-pxl), x-pxl : x+pxl]
        elif y+pxl> list.shape[0]: 
            paddingarr[ 0:2*pxl-(y+pxl-list.shape[0]), 0: 2*pxl]= list[y-pxl:list.shape[0],  x-pxl : x+pxl]
        else: 
            paddingarr[0:2*pxl, 0: 2*pxl]= list[y-pxl: y+pxl, x-pxl : x+pxl]    
    if paddingarr.shape[0] != 2*pxl or paddingarr.shape[1] !=2*pxl:
        print('for pxl= ' + str(2*pxl) +' paddingarr.shape not' + str(paddingarr.shape))
    return paddingarr
        
        
def overwritepara(dcmfile, arr, pixel,pxlspacing):
    dcmfile.PixelData = arr.tobytes()
    dcmfile.Rows=pixel
    dcmfile.Columns=pixel
    dcmfile.PixelSpacing=pxlspacing

def create_h5(i,ct_arr,ori,spacing,x1_arr,x2_arr):
    h5f = h5py.File(savepath + str(i+1).zfill(3) + '\\' +'ct_xray_data.h5', 'w')
    h5f.create_dataset('ct', data=ct_arr)
    h5f.create_dataset('ori_size', data=ori, shape=(), dtype='<i8')
    h5f.create_dataset('spacing', data=spacing, shape=(3,), dtype='<f8')
    h5f.create_dataset('xray1', data=x1_arr)
    h5f.create_dataset('xray2', data=x2_arr)
    h5f.close()

path= "*"
folders=os.listdir(path)

path_ct="\\preop\\CT\\DICOM\\DICOM\\"
path_EOS="\\preop\\EOS\\DICOM\\"
savepath= '*'

savecutEOS="*"

# Read ReferencePoint List
excelfile = load_workbook('*')
exceldata=excelfile.active

#Constant parameters for EOS images
S2D= 1300
S2I_y= 918
S2I_x= 987
Pxl_spacing= 0.179363

#Constant parameters for CT images
cubedim= 150 
x2ctpxl=256

#make directories for saving h5 files, if not existing
for i in range(0,402):
    foldername= str(i+1).zfill(3)
    if not os.path.exists(savepath + foldername):
        os.mkdir(savepath + foldername)

for i in range(0,201):
    #patients not to consider. Either missing data or ambiguous images
    if i==14 or i==39 or i==34 or i==46 or i==127 or i==159 or i==121 or i==139:
        continue

    #Reading EOS data
    data_f= dcmread(path + folders[i] + path_EOS + str(i+1).zfill(3) + '_preop_frontal.dcm')
    data_l= dcmread(path + folders[i] + path_EOS + str(i+1).zfill(3) + '_preop_lateral.dcm')
    
    FoV_f=data_f.FieldOfViewDimensions[1] 
    FoV_l=data_l.FieldOfViewDimensions[1]
    
    # Isocenter in the middle of the images
    iso_f= data_f.Columns /2
    iso_l= data_l.Columns /2
    
    #Read x-coordinates of referencepoints
    Rfr_f_left=float(exceldata.cell(i+2,2).value.split(';')[0]) 
    Rfr_f_right=float(exceldata.cell(i+2,3).value.split(';')[0])
    Rfr_l_left=float(exceldata.cell(i+2,4).value.split(';')[0])
    Rfr_l_right=float(exceldata.cell(i+2,5).value.split(';')[0])

    # distance between reference coordinate and isocenter
    d_l_right= (Rfr_l_right - iso_l) * Pxl_spacing
    d_l_left=  (Rfr_l_left - iso_l) * Pxl_spacing
    d_f_right= (Rfr_f_right - iso_f) * Pxl_spacing
    d_f_left=  (Rfr_f_left - iso_f) * Pxl_spacing
    
    # New width of the image 
    x_f_right = (FoV_f / S2I_x) * (S2I_x - abs(d_l_right))
    x_f_left = (FoV_f / S2I_x) * (S2I_x + abs(d_l_left))
    x_l_right = (FoV_l / S2I_y) * (S2I_y + abs(d_f_right))
    x_l_left = (FoV_l / S2I_y) * (S2I_y - abs(d_f_left))
    
    # New Pixelspacing
    pxlsp_f_right= Pxl_spacing * x_f_right / FoV_f
    pxlsp_f_left= Pxl_spacing * x_f_left / FoV_f
    pxlsp_l_right= Pxl_spacing * x_l_right / FoV_l
    pxlsp_l_left= Pxl_spacing * x_l_left / FoV_l
    
    #Resize image to the required pixel resolution at the left and right knee for resolution 15*15 cm
    pxl_f_right_cut= int(cubedim / pxlsp_f_right)
    pxl_f_left_cut= int(cubedim / pxlsp_f_left)
    pxl_l_right_cut= int(cubedim / pxlsp_l_right)
    pxl_l_left_cut= int(cubedim / pxlsp_l_left)
    
    #Reading pixel array 
    arr_f= data_f.pixel_array 
    arr_l= data_l.pixel_array 
    
    #Cut knees of the frontal EOS image
    CutArr_f_right = cutImg(arr_f, int(exceldata.cell(i+2,3).value.split(';')[0]), int(exceldata.cell(i+2,3).value.split(';')[1]), pxl_f_right_cut)
    CutArr_f_left = cutImg(arr_f, int(exceldata.cell(i+2,2).value.split(';')[0]), int(exceldata.cell(i+2,2).value.split(';')[1]), pxl_f_left_cut)
    
    #Cut knees of the lateral EOS image
    CutArr_l_right = cutImg(arr_l, int(exceldata.cell(i+2,5).value.split(';')[0]), int(exceldata.cell(i+2,5).value.split(';')[1]), pxl_l_right_cut)
    CutArr_l_left = cutImg(arr_l, int(exceldata.cell(i+2,4).value.split(';')[0]), int(exceldata.cell(i+2,4).value.split(';')[1]), pxl_l_left_cut)
        
    # Turns array into 256*256 shape 
    Arr_x2ct_f_right= resize(CutArr_f_right, (256, 256)) 
    Arr_x2ct_f_left= resize(CutArr_f_left, (256, 256)) 
    Arr_x2ct_l_right= resize(CutArr_l_right, (256, 256)) 
    Arr_x2ct_l_left= resize(CutArr_l_left, (256, 256)) 

    # Turn into unsigned 1 byte integer
    Arr_x2ct_f_right=img_as_ubyte(Arr_x2ct_f_right) 
    Arr_x2ct_f_left=img_as_ubyte(Arr_x2ct_f_left) 
    Arr_x2ct_l_right=img_as_ubyte(Arr_x2ct_l_right) 
    Arr_x2ct_l_left=img_as_ubyte(Arr_x2ct_l_left)
    
    # adapting the right endianese |u1
    Arr_x2ct_f_right =Arr_x2ct_f_right.astype('|u1')
    Arr_x2ct_f_left =Arr_x2ct_f_left.astype('|u1')
    Arr_x2ct_l_right =Arr_x2ct_l_right.astype('|u1')
    Arr_x2ct_l_left =Arr_x2ct_l_left.astype('|u1')

    #########################################################################
    # Reading CT data
    Files=os.listdir(path + str(i+1).zfill(3) + path_ct)
    #PixelSpacing isotrop and identical for 1 patient, but different over patients    
    data_ct= dcmread(path + folders[i] + path_ct + Files[0]) #reading dcm data from the first dcm file in a ct list
    pxlspacing= data_ct.PixelSpacing
    pixel= int(cubedim / pxlspacing[0])
    arr=data_ct.pixel_array

    # Calculating real distance in z-axis    
    data_ctlast= dcmread(path + folders[i] + path_ct + Files[-1]) 
    deltaZ= abs(data_ctlast.ImagePositionPatient[2] - data_ct.ImagePositionPatient[2])
    spacing_slices= (deltaZ + data_ct.SliceThickness) / len(Files)
    pixelheight= int(cubedim/ spacing_slices)
    
    if pixel % 2 == 1:
        pixel=pixel+1    
    
    if pixelheight % 2 == 1:
        pixelheight=pixelheight+1
    
    # Reference points right and left knee 
    refpoint_r=[int(i) for i in exceldata.cell(i+2,7).value.split(';')]
    refpoint_l=[int(i) for i in exceldata.cell(i+2,6).value.split(';')]
    
    # Height limits that shall be considered for cutting in z axis
    min_z_l= refpoint_l[2] - int(pixelheight/2)
    max_z_l= refpoint_l[2] + int(pixelheight/2) 
    min_z_r= refpoint_r[2] - int(pixelheight/2)
    max_z_r= refpoint_r[2] + int(pixelheight/2)    
    
    CTs3darr_right = np.zeros((pixel,pixel, pixelheight))
    CTs3darr_left = np.zeros((pixel,pixel, pixelheight))
    
    # z counter, to count filled slices in CTs3darr            
    z_counter_r=0
    z_counter_l=0
    
    for j in range(0, len(Files)-1):
        data_ct= dcmread(path + folders[i] + path_ct + Files[j])
        arr=data_ct.pixel_array 
        
        # Delete fragment influence and resize to [0, 2500]
        arr[arr>3000]=3000
        arr[arr<-1000]=-1000
        arr= arr +1000
        arr= arr*(2500/4000)
        
        #cut knee        
        arr_cut_r= cutImg(arr, refpoint_r[0], refpoint_r[1], pixel) #float64
        arr_cut_l= cutImg(arr, refpoint_l[0], refpoint_l[1], pixel) #float64
                
        # If cut height goes below CT boundaries
        if min_z_r <0:
            if min_z_r <= j < max_z_r:
                CTs3darr_right[:,:,abs(min_z_r)+j]= arr_cut_r
        if min_z_l <0:
            if min_z_l <= j < max_z_l:
                CTs3darr_left[:,:,abs(min_z_l)+j]= arr_cut_l
        
                # If cut height is within CT boundaries
        if min_z_r >=0:
            if min_z_r <= j < max_z_r:
                CTs3darr_right[:,:,z_counter_r]= arr_cut_r
                z_counter_r += 1
        
        if min_z_l >=0:
            if min_z_l <= j < max_z_l:
                CTs3darr_left[:,:,z_counter_l]= arr_cut_l
                z_counter_l +=1
   
    # Turns Array into 256*256*256 shape in float64 
    Arr_reszd_right = resize(CTs3darr_right, (256, 256, 256), preserve_range=True)
    Arr_reszd_left = resize(CTs3darr_left, (256, 256, 256), preserve_range=True)

    #adapting right endianese
    Arr_reszd_right = Arr_reszd_right.astype('<u2')
    Arr_reszd_left = Arr_reszd_left.astype('<u2')
    
    #Swap axis so they are identical with X2CT Dataset
    Arr_reszd_right = np.swapaxes(Arr_reszd_right,0,2)
    Arr_reszd_right = np.swapaxes(Arr_reszd_right,1,2)
    Arr_reszd_right = np.flip(Arr_reszd_right, axis=0)
    
    Arr_reszd_left = np.swapaxes(Arr_reszd_left,0,2)
    Arr_reszd_left = np.swapaxes(Arr_reszd_left,1,2)
    Arr_reszd_left = np.flip(Arr_reszd_left, axis=0)
    
    # Saving arrays for both knees into a h5 file in X2CT formats
    create_h5(i, Arr_reszd_right, 320, [1,1,1], Arr_x2ct_f_right, Arr_x2ct_l_right)
    create_h5(i+201, Arr_reszd_left, 320, [1,1,1], Arr_x2ct_f_left, Arr_x2ct_l_left)
    